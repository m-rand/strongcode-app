#!/usr/bin/env python3
"""
Convert StrongCode Excel program files to JSON format.

Approach: Parse the Results sheet as the primary data source.
Each lift block in Results contains:
  - Header: lift name, 1RM, NL, ARI, patterns
  - Variant definitions with color codes
  - Per-week rows: zone %1RM, weights (main + assist per variant), set×rep data
    distributed across session A/B/C columns, with cell colors indicating variant

Input sheet provides:
  - Client name, lift configs, session assignments
  - Volume patterns, session distribution settings

Usage:
    python3 scripts/excel_to_json.py "data/excel/2025-02 Katka SC60.xlsx"
    python3 scripts/excel_to_json.py --all
"""

import json
import os
import re
import sys
import unicodedata
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def slugify(name: str) -> str:
    nfkd = unicodedata.normalize('NFD', name)
    ascii_only = ''.join(c for c in nfkd if not unicodedata.combining(c))
    return re.sub(r'[^a-z0-9]+', '-', ascii_only.lower()).strip('-')


def col_to_num(col):
    if isinstance(col, str):
        return openpyxl.utils.column_index_from_string(col)
    return col


def safe_float(val, default=0.0):
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def safe_int(val, default=0):
    if val is None:
        return default
    try:
        return int(round(float(val)))
    except (ValueError, TypeError):
        return default


def cell_val(ws, row, col):
    return ws.cell(row=row, column=col_to_num(col)).value


# Blue color normalization: variant definitions use FFCFE2F3,
# but set cells sometimes use FFC9DAF8 (different shade of blue).
BLUE_COLORS = {'FFCFE2F3', 'FFC9DAF8'}
CANONICAL_BLUE = 'FFCFE2F3'


def cell_color(ws, row, col):
    c = ws.cell(row=row, column=col_to_num(col))
    f = c.fill
    if f and f.start_color and f.start_color.rgb:
        rgb = f.start_color.rgb
        if rgb not in ('00000000', '0'):
            # Normalize blue shades to a single canonical value
            if rgb in BLUE_COLORS:
                return CANONICAL_BLUE
            return rgb
    return None


LIFT_NAMES = {
    'squat': 'squat',
    'bench press': 'bench_press',
    'benchpress': 'bench_press',
    'bench': 'bench_press',
    'deadlift': 'deadlift',
}


def normalize_lift(name):
    if name is None:
        return None
    return LIFT_NAMES.get(name.lower().strip(), name.lower().strip())


# ---------------------------------------------------------------------------
# Session columns in Results sheet
# ---------------------------------------------------------------------------

# Each session block has 10 set columns + 1 total column
# Session A: N(14) to W(23), total at X(24)
# Session B: Y(25) to AH(34), total at AI(35)
# Session C: AJ(36) to AS(45), total at AT(46)
# Remaining/check column: AU(47)

SESSION_CONFIGS = [
    {
        'name': 'A',
        'start': col_to_num('N'),  # 14
        'end': col_to_num('W'),    # 23
        'total': col_to_num('X'),  # 24
    },
    {
        'name': 'B',
        'start': col_to_num('Y'),  # 25
        'end': col_to_num('AH'),   # 34
        'total': col_to_num('AI'), # 35
    },
    {
        'name': 'C',
        'start': col_to_num('AJ'), # 36
        'end': col_to_num('AS'),   # 45
        'total': col_to_num('AT'), # 46
    },
]

# Weight columns for variants:  D=main, H=assist1, J=assist2, L=assist3
WEIGHT_COLS = {
    'main': col_to_num('D'),   # 4
    'assist1': col_to_num('H'),  # 8
    'assist2': col_to_num('J'),  # 10
    'assist3': col_to_num('L'),  # 12
}


# ---------------------------------------------------------------------------
# Results Sheet Parser
# ---------------------------------------------------------------------------

def find_lift_blocks(ws):
    """Find row numbers where each lift block starts (header row with 'Name' + 'lift')."""
    blocks = []
    for r in range(1, ws.max_row + 1):
        if cell_val(ws, r, 'A') == 'Name' and cell_val(ws, r, 'G') == 'lift':
            blocks.append(r)
    return blocks


def parse_variants(ws_data, ws_style, start_row, end_row):
    """
    Extract variant definitions with colors from the Notes section.
    Variants appear near the top of each lift block in column E.
    The first variant (no background color) is the competition variant.
    Subsequent variants have background colors (yellow, pink, blue).
    """
    variants = []

    # Scan rows after header for variant names in column E
    for r in range(start_row + 2, min(start_row + 12, end_row)):
        e_val = cell_val(ws_data, r, 'E')
        if e_val is None:
            continue

        name = str(e_val).strip()
        # Skip known non-variant labels
        if name.lower() in ('#reps', 'rep rng', '%1rm', 'rpe', 'notes', ''):
            continue
        # Skip if it looks like a number
        try:
            float(name)
            continue
        except ValueError:
            pass

        e_color = cell_color(ws_style, r, 'E')
        variants.append({
            'name': name,
            'color': e_color,
        })

    # Assign weight columns based on cell background color.
    # Colors map consistently to specific columns regardless of lift:
    #   No color (None)       -> col D (main weight)
    #   Yellow (FFFFF2CC)     -> col H (assist1)
    #   Pink   (FFF4CCCC)     -> col J (assist2)
    #   Blue   (FFCFE2F3)     -> col L (assist3)
    COLOR_TO_WEIGHT_COL = {
        None:         'main',
        'FFFFF2CC':   'assist1',  # yellow
        'FFF4CCCC':   'assist2',  # pink
        'FFCFE2F3':   'assist3',  # blue
    }
    for v in variants:
        v['weight_col'] = COLOR_TO_WEIGHT_COL.get(v['color'], 'main')

    return variants


def find_data_start_row(ws, header_row, end_row):
    """
    Find the first data row in a lift block.
    Data rows start when column A has a week number (1-8).
    """
    for r in range(header_row + 5, min(header_row + 20, end_row)):
        a_val = cell_val(ws, r, 'A')
        if a_val is not None:
            try:
                wk = int(round(float(a_val)))
                if 1 <= wk <= 8:
                    return r
            except (ValueError, TypeError):
                pass
    return None


def parse_lift_block(ws_data, ws_style, header_row, end_row):
    """Parse a single lift block from the Results sheet."""
    info_row = header_row + 1

    lift_name = normalize_lift(cell_val(ws_data, info_row, 'G'))
    one_rm = safe_float(cell_val(ws_data, info_row, 'N'))
    rounding = safe_float(cell_val(ws_data, info_row, 'P'), 2.5)
    block = cell_val(ws_data, info_row, 'R') or 'prep'
    total_nl = safe_int(cell_val(ws_data, info_row, 'T'))
    ari = safe_float(cell_val(ws_data, info_row, 'Y'))
    pattern_main = cell_val(ws_data, info_row, 'AA')
    pattern_8190 = cell_val(ws_data, info_row, 'AC')

    nl_90 = safe_int(cell_val(ws_data, info_row, 'V'))
    if nl_90 == 0:
        nl_90 = safe_int(cell_val(ws_data, info_row, 'W'))

    if not lift_name:
        return None

    # Parse variant definitions
    variants = parse_variants(ws_data, ws_style, header_row, end_row)

    # Build color -> variant mapping
    color_to_variant = {}
    for v in variants:
        if v['color']:
            color_to_variant[v['color']] = v
    # Default (no color) variant
    default_variant = next(
        (v for v in variants if v['color'] is None),
        {'name': lift_name, 'weight_col': 'main', 'color': None}
    )

    # Find where data rows start
    data_start = find_data_start_row(ws_data, header_row, end_row)
    if data_start is None:
        return {
            'lift': lift_name, 'one_rm': one_rm, 'rounding': rounding,
            'block': block, 'total_nl': total_nl, 'ari': ari, 'nl_90': nl_90,
            'pattern_main': str(pattern_main) if pattern_main else None,
            'pattern_8190': str(pattern_8190) if pattern_8190 else None,
            'variants': variants, 'weeks': {},
        }

    # Parse data rows
    week_data = {}
    current_week = None

    for r in range(data_start, end_row + 1):
        a_val = cell_val(ws_data, r, 'A')
        c_val = cell_val(ws_data, r, 'C')  # zone %1RM

        # Detect new week
        if a_val is not None:
            try:
                wk = int(round(float(a_val)))
                if 1 <= wk <= 8:
                    current_week = wk
                    if current_week not in week_data:
                        week_data[current_week] = []
            except (ValueError, TypeError):
                pass

        if current_week is None:
            continue

        # Skip RPE rows
        if c_val is not None and str(c_val).strip().upper() == 'RPE':
            continue

        # Only process rows with valid zone percentages in column C
        if c_val is None:
            continue
        zone_pct = safe_float(c_val)
        if zone_pct < 0.1 or zone_pct > 1.0:
            continue

        # Read weights for each variant column
        weights = {}
        for key, col_num in WEIGHT_COLS.items():
            w = safe_float(cell_val(ws_data, r, col_num))
            if w > 0:
                weights[key] = w

        # Read reps value from column E
        base_reps = safe_int(cell_val(ws_data, r, 'E'))

        # Read sets for each session, tracking column positions
        # so we can later sort by column to preserve interleaving across zones
        session_sets = {}
        for sess_cfg in SESSION_CONFIGS:
            sets = []
            for col in range(sess_cfg['start'], sess_cfg['end'] + 1):
                reps_val = cell_val(ws_data, r, col)
                if reps_val is not None and safe_float(reps_val) > 0:
                    reps = safe_int(reps_val)
                    clr = cell_color(ws_style, r, col)

                    # Determine variant from cell color
                    if clr and clr in color_to_variant:
                        variant = color_to_variant[clr]
                    else:
                        variant = default_variant

                    # Get weight for this variant
                    weight = weights.get(variant['weight_col'], weights.get('main', 0))

                    sets.append({
                        'reps': reps,
                        'weight': weight,
                        'variant': variant['name'],
                        'zone_pct': zone_pct,
                        'col_pos': col,  # track position for interleaving
                    })

            if sets:
                session_sets[sess_cfg['name']] = sets

        if session_sets:
            week_data[current_week].append({
                'zone_pct': zone_pct,
                'weights': weights,
                'session_sets': session_sets,
            })

    return {
        'lift': lift_name,
        'one_rm': one_rm,
        'rounding': rounding,
        'block': block,
        'total_nl': total_nl,
        'ari': ari,
        'nl_90': nl_90,
        'pattern_main': str(pattern_main) if pattern_main else None,
        'pattern_8190': str(pattern_8190) if pattern_8190 else None,
        'variants': variants,
        'weeks': week_data,
    }


def parse_results_sheet(ws_data, ws_style):
    """Parse the entire Results sheet, returning data for all lifts."""
    block_rows = find_lift_blocks(ws_data)
    lifts = []

    for i, header_row in enumerate(block_rows):
        end_row = block_rows[i + 1] - 1 if i + 1 < len(block_rows) else ws_data.max_row
        lift = parse_lift_block(ws_data, ws_style, header_row, end_row)
        if lift:
            lifts.append(lift)
            print(f"  Parsed {lift['lift']}: {len(lift['variants'])} variants, "
                  f"{len(lift['weeks'])} weeks, 1RM={lift['one_rm']}")
            for v in lift['variants']:
                print(f"    Variant: {v['name']} (color={v['color']}, weight_col={v['weight_col']})")

    return lifts


# ---------------------------------------------------------------------------
# Input Sheet Parser (for supplementary data)
# ---------------------------------------------------------------------------

def parse_input_sheet(ws):
    """Parse Input sheet for client name, lift configs, and session assignments."""
    # Client name from first 'Name' row
    client_name = "Unknown"
    for r in range(1, 10):
        if cell_val(ws, r, 'A') == 'Name' and cell_val(ws, r, 'B'):
            client_name = str(cell_val(ws, r, 'B')).strip()
            break

    # Find all lift input blocks
    lift_configs = {}
    lift_start_rows = []
    for r in range(1, ws.max_row + 1):
        if cell_val(ws, r, 'A') == 'Name' and cell_val(ws, r, 'B'):
            lift_start_rows.append(r)

    for i, start_row in enumerate(lift_start_rows):
        end_row = lift_start_rows[i + 1] - 1 if i + 1 < len(lift_start_rows) else ws.max_row

        # Lift name is in B of the row after "Name"
        lift_name = normalize_lift(cell_val(ws, start_row + 1, 'B'))
        if not lift_name:
            continue

        config = {'skill_level': cell_val(ws, start_row + 1, 'C') or 'intermediate'}

        # Zone data
        zone_map = {
            '50-60%': '55', '61-70%': '65', '71-80%': '75',
            '81-90%': '85', '91-94%': '90', '95-100%': '95'
        }
        zones = {}
        for r in range(start_row + 8, min(start_row + 20, end_row)):
            label = cell_val(ws, r, 'A')
            if label and str(label) in zone_map:
                zk = zone_map[str(label)]
                zones[zk] = {
                    'weight': safe_float(cell_val(ws, r, 'B')),
                    'percentage': safe_float(cell_val(ws, r, 'D')),
                    'reps': safe_int(cell_val(ws, r, 'E')),
                }
        config['zones'] = zones

        # Volume, weeks
        config['volume'] = safe_int(cell_val(ws, start_row + 5, 'B'))
        config['weeks'] = safe_int(cell_val(ws, start_row + 6, 'B'), 4)

        # Session distribution — take the FIRST match (there may be
        # multiple options listed, e.g. 3 days + 2 days alternatives)
        for r in range(start_row + 28, min(start_row + 50, end_row)):
            b_val = cell_val(ws, r, 'B')
            if b_val and 'days / week' in str(b_val) and 'sessions_per_week' not in config:
                config['sessions_per_week'] = safe_int(cell_val(ws, r, 'A'))
            if b_val and 'distribution in a week' in str(b_val) and 'session_distribution' not in config:
                config['session_distribution'] = str(cell_val(ws, r, 'A'))

        # Weekly zone breakdown
        weekly_zones = {}
        for r in range(start_row + 20, min(start_row + 40, end_row)):
            if cell_val(ws, r, 'A') == 'Int.zone / Week':
                for zr in range(r + 2, r + 8):
                    zpct = safe_float(cell_val(ws, zr, 'A'))
                    zk = {
                        0.55: '55', 0.65: '65', 0.75: '75',
                        0.85: '85', 0.925: '90', 0.95: '95'
                    }.get(zpct)
                    if zk:
                        weekly_zones[zk] = {
                            f'week_{w}': safe_int(cell_val(ws, zr, w + 1))
                            for w in range(1, 5)
                        }
                break
        config['weekly_zones'] = weekly_zones

        lift_configs[lift_name] = config

    # Session assignments table
    # Structure: header row ("Week", "session 1", "session 2", ...)
    # Then data rows grouped per week — week number only on the first row,
    # subsequent rows (one per lift) have None in col A.
    # Example:
    #   1.0  | SQ A | BP B | SQ B | SQ C | --
    #   None | BP A | --   | DL B | BP C | --
    #   None | DL A | --   | --   | DL C | --
    session_assignments = {}
    for r in range(1, ws.max_row + 1):
        if cell_val(ws, r, 'A') == 'Week' and cell_val(ws, r, 'B') and 'session' in str(cell_val(ws, r, 'B')).lower():
            current_week = None
            for dr in range(r + 1, r + 30):
                a_val = cell_val(ws, dr, 'A')

                # New week starts when col A has a number
                if a_val is not None:
                    wk = safe_int(a_val)
                    if wk > 0:
                        current_week = wk

                # Skip if we haven't found a week yet
                if current_week is None:
                    continue

                if current_week not in session_assignments:
                    session_assignments[current_week] = {}

                # Check if this row has any data in session columns
                row_has_data = False
                for col in range(2, 7):  # B-F = session 1-5
                    val = cell_val(ws, dr, col)
                    if val and str(val).strip() != '--':
                        row_has_data = True
                        sess_num = col - 1
                        if sess_num not in session_assignments[current_week]:
                            session_assignments[current_week][sess_num] = []
                        session_assignments[current_week][sess_num].append(str(val).strip())

                # If row is completely empty (no week, no data), we're past the table
                if a_val is None and not row_has_data:
                    break
            break

    return client_name, lift_configs, session_assignments


# ---------------------------------------------------------------------------
# Build Sessions JSON from Results data
# ---------------------------------------------------------------------------

def build_sessions(results_lifts):
    """
    Build the sessions JSON structure from Results sheet data.

    Results gives us per-lift, per-week, per-session set data.
    We reorganize into: sessions[session_letter][week_N] = {lifts: [{lift, variant, sets}]}

    IMPORTANT: Sets are sorted by column position to preserve the original
    interleaving across zones (e.g. 3@75%, 1@85%, 3@75%, 1@85%...).
    """
    sessions = {}

    for lift_data in results_lifts:
        lift_name = lift_data['lift']

        for week_num, zone_rows in lift_data['weeks'].items():
            week_key = f'week_{week_num}'

            for sess_letter in ('A', 'B', 'C'):
                # Collect ALL sets for this lift/week/session across all zone rows
                all_sets = []
                for zone_row in zone_rows:
                    if sess_letter in zone_row['session_sets']:
                        all_sets.extend(zone_row['session_sets'][sess_letter])

                if not all_sets:
                    continue

                # Sort by column position to preserve interleaving
                all_sets.sort(key=lambda s: s['col_pos'])

                if sess_letter not in sessions:
                    sessions[sess_letter] = {}
                if week_key not in sessions[sess_letter]:
                    sessions[sess_letter][week_key] = {'lifts': []}

                # Build ordered list of sets, preserving the interleaved order.
                # We output sets as a flat ordered list for this lift.
                ordered_sets = []
                for s in all_sets:
                    ordered_sets.append({
                        'weight': s['weight'],
                        'reps': s['reps'],
                        'zone_pct': s['zone_pct'],
                        'variant': s['variant'],
                    })

                sessions[sess_letter][week_key]['lifts'].append({
                    'lift': lift_name,
                    'sets': ordered_sets,
                })

    return sessions


# ---------------------------------------------------------------------------
# Build Input & Calculated sections
# ---------------------------------------------------------------------------

def build_input_section(lift_configs, results_lifts):
    """Build the input section from Input sheet configs + Results metadata."""
    input_section = {}
    for rl in results_lifts:
        ln = rl['lift']
        cfg = lift_configs.get(ln, {})
        zones = cfg.get('zones', {})
        pct_75 = zones.get('75', {}).get('percentage', 0)
        pct_85 = zones.get('85', {}).get('percentage', 0)
        reps_90 = zones.get('90', {}).get('reps', 0)
        reps_95 = zones.get('95', {}).get('reps', 0)

        input_section[ln] = {
            'volume': cfg.get('volume', rl['total_nl']),
            'rounding': rl['rounding'],
            'weights': {
                z: zones[z]['weight']
                for z in ('55', '65', '75', '85', '90', '95')
                if z in zones and zones[z]['weight'] > 0
            },
            'intensity_distribution': {
                '75_percent': round(pct_75 * 100, 1) if 0 < pct_75 < 1 else round(pct_75, 1),
                '85_percent': round(pct_85 * 100, 1) if 0 < pct_85 < 1 else round(pct_85, 1),
                '90_total_reps': reps_90,
                '95_total_reps': reps_95,
            },
            'volume_pattern_main': rl['pattern_main'],
            'volume_pattern_8190': rl['pattern_8190'],
            'sessions_per_week': cfg.get('sessions_per_week', 3),
            'session_distribution': cfg.get('session_distribution'),
        }
    return input_section


def build_calculated(lift_configs, results_lifts):
    """Build the calculated section from Input weekly zone data."""
    zone_pct_map = {
        '55': 0.55, '65': 0.65, '75': 0.75,
        '85': 0.85, '90': 0.925, '95': 0.95,
    }
    calculated = {}

    for rl in results_lifts:
        ln = rl['lift']
        cfg = lift_configs.get(ln, {})
        wz = cfg.get('weekly_zones', {})
        zones_cfg = cfg.get('zones', {})

        zone_totals = {}
        for zk in ('55', '65', '75', '85', '90', '95'):
            total = sum(wz.get(zk, {}).get(f'week_{w}', 0) for w in range(1, 5))
            zone_totals[zk] = total  # include 55 even if 0

        total_nl = sum(zone_totals.values())
        weighted = sum(zone_totals.get(z, 0) * zone_pct_map.get(z, 0) for z in zone_totals)
        # Use ARI from Results header if available, otherwise calculate
        block_ari = round(rl['ari'] * 100, 2) if rl['ari'] > 0 else (
            round(weighted / total_nl * 100, 2) if total_nl > 0 else 0
        )

        calc = {
            '_summary': {
                'total_nl': cfg.get('volume', total_nl),
                'actual_nl': total_nl,
                'block_ari': block_ari,
                'zone_totals': zone_totals,
                'weights': {z: zones_cfg[z]['weight'] for z in zones_cfg if zones_cfg[z]['weight'] > 0},
            }
        }

        for week_num in range(1, 5):
            wk = f'week_{week_num}'
            zones = {
                zk: wz.get(zk, {}).get(wk, 0)
                for zk in ('55', '65', '75', '85', '90', '95')
            }
            wt = sum(zones.values())
            ww = sum(zones.get(z, 0) * zone_pct_map.get(z, 0) for z in zones)
            calc[wk] = {
                'total_reps': wt,
                'zones': zones,
                'ari': round(ww / wt * 100, 1) if wt > 0 else 0,
            }

        calculated[ln] = calc

    return calculated


# ---------------------------------------------------------------------------
# Main Conversion
# ---------------------------------------------------------------------------

def convert_excel_to_json(excel_path: str, output_dir: str = None) -> str:
    excel_path = Path(excel_path)
    if not excel_path.exists():
        raise FileNotFoundError(f"File not found: {excel_path}")

    print(f"\n{'='*60}")
    print(f"Converting: {excel_path.name}")
    print(f"{'='*60}")

    wb_data = openpyxl.load_workbook(str(excel_path), data_only=True)
    wb_style = openpyxl.load_workbook(str(excel_path))

    # 1) Parse Input sheet
    client_name, lift_configs, session_assignments = parse_input_sheet(wb_data['Input'])
    print(f"\n  Client: {client_name}")
    print(f"  Lift configs: {list(lift_configs.keys())}")
    print(f"  Session assignments: {session_assignments}")

    # 2) Parse Results sheet (primary data source)
    print(f"\n  Parsing Results sheet...")
    results_lifts = parse_results_sheet(wb_data['Results'], wb_style['Results'])

    # 3) Build sessions from Results data
    sessions = build_sessions(results_lifts)
    print(f"\n  Built sessions: {sorted(sessions.keys())}")
    for sl in sorted(sessions.keys()):
        for wk in sorted(sessions[sl].keys()):
            lifts = sessions[sl][wk].get('lifts', [])
            total_sets = sum(len(l['sets']) for l in lifts)
            # Count variants per lift
            summaries = []
            for l in lifts:
                variants_in_sets = set(s.get('variant', '?') for s in l['sets'])
                summaries.append(f"{l['lift'][:2].upper()}:{','.join(v[:10] for v in variants_in_sets)}({len(l['sets'])})")
            print(f"    {sl}/{wk}: {total_sets} sets - {', '.join(summaries)}")

    # 4) Build metadata
    block = results_lifts[0]['block'] if results_lifts else 'prep'
    skill = lift_configs.get(results_lifts[0]['lift'], {}).get('skill_level', 'intermediate') if results_lifts else 'intermediate'
    one_rm = {rl['lift']: rl['one_rm'] for rl in results_lifts}

    date_match = re.search(r'(\d{4})-(\d{2})', excel_path.stem)
    start_date = f"{date_match.group(1)}-{date_match.group(2)}-01" if date_match else datetime.now().strftime('%Y-%m-%d')
    weeks = lift_configs.get(results_lifts[0]['lift'], {}).get('weeks', 4) if results_lifts else 4

    client_slug = slugify(client_name)
    filename = f"{start_date}_{client_slug}_{block}_all_lifts.json"

    try:
        sd = datetime.strptime(start_date, '%Y-%m-%d')
        end_date = (sd + timedelta(weeks=weeks) - timedelta(days=1)).strftime('%Y-%m-%d')
    except Exception:
        end_date = None

    # 5) Assemble complete program JSON
    program = {
        'schema_version': '1.0',
        'meta': {
            'filename': filename,
            'created_at': datetime.now().isoformat() + 'Z',
            'created_by': 'Excel import',
            'status': 'active',
            'notes': f'Converted from {excel_path.name}',
            'source_file': excel_path.name,
        },
        'client': {
            'name': client_name,
            'delta': skill,
            'one_rm': one_rm,
        },
        'program_info': {
            'block': block,
            'start_date': start_date,
            'end_date': end_date,
            'weeks': weeks,
        },
        'input': build_input_section(lift_configs, results_lifts),
        'calculated': build_calculated(lift_configs, results_lifts),
        'sessions': sessions,
        'session_assignments': session_assignments,
    }

    # 6) Write output
    if output_dir is None:
        output_dir = str(excel_path.parent.parent / 'programs_json')
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, filename)

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(program, f, ensure_ascii=False, indent=2)

    total_sets = sum(
        len(e['sets'])
        for s in sessions.values()
        for w in s.values()
        for e in w.get('lifts', [])
    )
    print(f"\n  Written: {output_path}")
    print(f"     {len(sessions)} sessions, {total_sets} total set prescriptions")

    wb_data.close()
    wb_style.close()
    return output_path


def main():
    if len(sys.argv) < 2:
        print('Usage: python3 scripts/excel_to_json.py "path/to/file.xlsx" [output_dir]')
        print('       python3 scripts/excel_to_json.py --all')
        sys.exit(1)

    if sys.argv[1] == '--all':
        excel_dir = Path('data/excel')
        files = sorted(excel_dir.glob('*.xlsx'))
        if not files:
            print("No .xlsx files found in data/excel/")
            sys.exit(1)
        print(f"Found {len(files)} Excel files")
        output_dir = str(Path('data/programs_json'))
        for f in files:
            try:
                convert_excel_to_json(str(f), output_dir)
            except Exception as e:
                print(f"\n  Error: {f.name}: {e}")
                import traceback
                traceback.print_exc()
    else:
        convert_excel_to_json(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else None)


if __name__ == '__main__':
    main()
